import os
import pandas as pd
import numpy as np
from flask import Flask, render_template, request, send_file, flash, redirect, url_for, session
from werkzeug.utils import secure_filename
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pickle
import uuid
import gc  # For garbage collection

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-this-to-random-string'  # Change this to a random secret key

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB max upload size
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
CACHE_FOLDER = 'cache'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)
os.makedirs(CACHE_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER
app.config['CACHE_FOLDER'] = CACHE_FOLDER

# Helper functions for session data storage using file-based cache
def save_to_session(df):
    """Save DataFrame to file-based cache and store reference in session"""
    # Generate unique session ID if not exists
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    
    cache_file = os.path.join(app.config['CACHE_FOLDER'], f"{session['session_id']}.pkl")
    with open(cache_file, 'wb') as f:
        pickle.dump(df, f)
    
    session['has_data'] = True

def load_from_session():
    """Load DataFrame from file-based cache"""
    if 'session_id' in session and session.get('has_data'):
        cache_file = os.path.join(app.config['CACHE_FOLDER'], f"{session['session_id']}.pkl")
        if os.path.exists(cache_file):
            with open(cache_file, 'rb') as f:
                return pickle.load(f)
    return None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_gstin(gstin_name_string):
    """Extract GSTIN from the format '01AAACI6306G1Z7 / IND LABORATORIES LTD'"""
    if pd.isna(gstin_name_string) or not isinstance(gstin_name_string, str):
        return None
    
    parts = gstin_name_string.split('/')
    if len(parts) > 0:
        return parts[0].strip()
    return None

def extract_pan(gstin):
    """Extract PAN from GSTIN (characters 3-10, both included)"""
    if pd.isna(gstin) or not isinstance(gstin, str) or len(gstin) < 10:
        return None

    return gstin[2:12]  # Python uses 0-based indexing, so 2:10 gives characters 3-10 (both included)

def extract_name(gstin_name_string):
    """Extract name from the format '01AAACI6306G1Z7 / IND LABORATORIES LTD'"""
    if pd.isna(gstin_name_string) or not isinstance(gstin_name_string, str):
        return None
    
    parts = gstin_name_string.split('/')
    if len(parts) > 1:
        return parts[1].strip()
    return None

def extract_date(ewb_date_string):
    """Extract and format date from '6713 - 02/09/2023 17:41:00' to '02-09-2023'"""
    if pd.isna(ewb_date_string) or not isinstance(ewb_date_string, str):
        return None
    
    # Split by ' - ' to get the date part
    parts = ewb_date_string.split(' - ')
    if len(parts) > 1:
        date_time_part = parts[1].strip()
        # Extract just the date part (before the time)
        date_part = date_time_part.split(' ')[0]
        # Convert from MM/DD/YYYY to DD-MM-YYYY format
        try:
            date_obj = datetime.strptime(date_part, '%d/%m/%Y')
            return date_obj.strftime('%d-%m-%Y')
        except ValueError:
            try:
                # Try MM/DD/YYYY format
                date_obj = datetime.strptime(date_part, '%m/%d/%Y')
                return date_obj.strftime('%d-%m-%Y')
            except ValueError:
                return date_part.replace('/', '-')
    return None

def extract_pan_from_gstin_name(gstin_name_string):
    """Extract PAN from From GSTIN & Name for duplicate checking"""
    if pd.isna(gstin_name_string) or not isinstance(gstin_name_string, str):
        return None
    
    gstin = extract_gstin(gstin_name_string)
    return extract_pan(gstin)

def round_assess_value(assess_val):
    """Round the assess value similar to Excel ROUND function with -4"""
    if pd.isna(assess_val):
        return None
    
    try:
        # Convert to float if it's not already
        val = float(assess_val)
        # Round to nearest 10000 (equivalent to ROUND(value, -4))
        return round(val, -4)
    except (ValueError, TypeError):
        return assess_val

def process_excel_file(input_path, output_path):
    """Process the Excel file according to the specifications"""
    try:
        # Read the Excel file with engine openpyxl for better performance
        df = pd.read_excel(input_path, engine='openpyxl')
        
        # Create a copy for processing
        processed_df = df.copy()
        
        # Clear original df to free memory
        del df
        gc.collect()
        
        # 1. Change 'EWB No.' to 'Serial No'
        if 'EWB No.' in processed_df.columns:
            processed_df.rename(columns={'EWB No.': 'Serial No'}, inplace=True)
        
        # 2. Rename 'EWB No. & Dt.' to 'Serial No. & Dt.'
        if 'EWB No. & Dt.' in processed_df.columns:
            processed_df.rename(columns={'EWB No. & Dt.': 'Serial No. & Dt.'}, inplace=True)
        
        # 3. Extract GSTIN, PAN, and NAME from 'To GSTIN & Name'
        if 'To GSTIN & Name' in processed_df.columns:
            processed_df['GSTIN'] = processed_df['To GSTIN & Name'].apply(extract_gstin)
            processed_df['PAN'] = processed_df['GSTIN'].apply(extract_pan)
            processed_df['NAME'] = processed_df['To GSTIN & Name'].apply(extract_name)
        
        # Handle case where 'Name' column already exists (case sensitive)
        if 'Name' in processed_df.columns and 'NAME' not in processed_df.columns:
            processed_df['NAME'] = processed_df['Name']
        elif 'NAME' in processed_df.columns and 'Name' not in processed_df.columns:
            # NAME column already exists, keep it as is
            pass
        
        # Ensure NAME column is created properly
        if 'NAME' not in processed_df.columns or processed_df['NAME'].isnull().all():
            print("Warning: NAME column is missing or empty. Check 'To GSTIN & Name' or 'Name' data.")
            processed_df['NAME'] = 'Unknown'  # Fallback value for missing names
        
        # 4. Extract Date from 'Serial No. & Dt.'
        if 'Serial No. & Dt.' in processed_df.columns:
            processed_df['Date'] = processed_df['Serial No. & Dt.'].apply(extract_date)
        
        # 5. Create VALUE column by rounding 'Assess Val.'
        if 'Assess Val.' in processed_df.columns:
            processed_df['VALUE'] = processed_df['Assess Val.'].apply(round_assess_value)
        
        # 6. Remove duplicate rows based on PAN appearing in 'From GSTIN & Name'
        if 'From GSTIN & Name' in processed_df.columns and 'PAN' in processed_df.columns:
            # Extract PAN from 'From GSTIN & Name' for comparison
            processed_df['From_PAN'] = processed_df['From GSTIN & Name'].apply(extract_pan_from_gstin_name)
            
            # Remove rows where PAN (from To GSTIN) equals From_PAN (from From GSTIN)
            before_count = len(processed_df)
            processed_df = processed_df[processed_df['PAN'] != processed_df['From_PAN']]
            after_count = len(processed_df)
            duplicates_removed = before_count - after_count
            
            # Drop the temporary From_PAN column
            processed_df.drop(columns=['From_PAN'], inplace=True)
            
            print(f"Removed {duplicates_removed} duplicate rows based on PAN matching")
        
        # 6.5. Remove duplicate rows from the entire dataset
        before_dedup = len(processed_df)
        processed_df = processed_df.drop_duplicates()
        after_dedup = len(processed_df)
        total_duplicates_removed = before_dedup - after_dedup
        print(f"Removed {total_duplicates_removed} duplicate rows from processed data")
        
        # 7. Delete specified columns after creating new ones
        columns_to_delete = ['To GSTIN & Name', 'Assess Val.', 'Tax Val.', 'Latest Vehicle No.']
        # Don't delete 'Name' column if it exists and we're using it for NAME
        if 'Name' in processed_df.columns and processed_df['NAME'].equals(processed_df['Name']):
            # Keep the original 'Name' column, but we can optionally rename it to avoid confusion
            pass
        for col in columns_to_delete:
            if col in processed_df.columns:
                processed_df.drop(columns=[col], inplace=True)
        
        # 8. Reorder columns according to specification
        desired_order = [
            'Serial No',        # Was EWB No., now renamed to Serial No
            'From GSTIN & Name', 
            'GSTIN',
            'PAN',
            'NAME',
            'From Place & Pin',
            'To Place & Pin',
            'Serial No. & Dt.',  # Was EWB No. & Dt., now renamed to Serial No. & Dt.
            'Date',
            'Doc No. & Dt.',
            'VALUE',
            'HSN Code',
            'HSN Desc.'
        ]
        
        # Only include columns that exist in the dataframe
        existing_columns = [col for col in desired_order if col in processed_df.columns]
        
        # Add any remaining columns that weren't in the desired order
        remaining_columns = [col for col in processed_df.columns if col not in existing_columns]
        final_column_order = existing_columns + remaining_columns
        
        # Reorder the dataframe
        processed_df = processed_df[final_column_order]
        
        # Save the processed file
        processed_df.to_excel(output_path, index=False)
        
        return True, "File processed successfully!"
        
    except Exception as e:
        return False, f"Error processing file: {str(e)}"

def extract_seller_name(from_gstin_name_string):
    """Extract seller company name from 'From GSTIN & Name' field (after /)"""
    if pd.isna(from_gstin_name_string) or not isinstance(from_gstin_name_string, str):
        return None
    
    parts = from_gstin_name_string.split('/')
    if len(parts) > 1:
        return parts[1].strip()
    return None

def extract_seller_gstin(from_gstin_name_string):
    """Extract seller GSTIN from 'From GSTIN & Name' field (before /)"""
    if pd.isna(from_gstin_name_string) or not isinstance(from_gstin_name_string, str):
        return None
    
    parts = from_gstin_name_string.split('/')
    if len(parts) > 0:
        return parts[0].strip()
    return None

def extract_seller_pan(from_gstin_name_string):
    """Extract seller PAN from 'From GSTIN & Name' field"""
    gstin = extract_seller_gstin(from_gstin_name_string)
    return extract_pan(gstin)

def generate_summary(processed_df):
    """Generate summary grouped by PAN and NAME with product details"""
    try:
        # Group by PAN and NAME, then aggregate unique HSN Desc. and sum VALUE
        summary = processed_df.groupby(['PAN', 'NAME', 'HSN Desc.']).agg(
            product_value=('VALUE', 'sum')
        ).reset_index()

        # Add Grand Total row for each PAN
        grand_totals = summary.groupby(['PAN', 'NAME']).agg(
            total_value=('product_value', 'sum')
        ).reset_index()

        # Merge product details with grand totals
        summary = pd.merge(summary, grand_totals, on=['PAN', 'NAME'], how='left')

        return summary
    except Exception as e:
        print(f"Error generating summary: {str(e)}")
        return None

def generate_seller_analysis(processed_df):
    """Generate competitive analysis summary by seller companies"""
    try:
        # First, extract seller names and PANs from 'From GSTIN & Name'
        processed_df['SELLER_NAME'] = processed_df['From GSTIN & Name'].apply(extract_seller_name)
        processed_df['SELLER_PAN'] = processed_df['From GSTIN & Name'].apply(extract_seller_pan)
        
        # Calculate quantity if 2024-25 price column exists
        # Quantity (MT) = Assess Val. / (2024-25 * 1000)
        
        # Print available columns for debugging
        print(f"\n=== SELLER ANALYSIS DEBUG ===")
        print(f"Available columns: {list(processed_df.columns)}")
        
        # Check if QTY.MT already exists (from Data Cleaner)
        if 'QTY.MT' in processed_df.columns:
            print("Using pre-calculated QTY.MT column from Data Cleaner")
            processed_df['QUANTITY_MT'] = processed_df['QTY.MT']
            print(f"Sample quantities: {processed_df['QUANTITY_MT'].head().tolist()}")
            print(f"=== END DEBUG ===\n")
        else:
            # Calculate it dynamically if not pre-calculated
            print("QTY.MT not found, calculating dynamically...")
            
            # Find column that contains '2024-25'
            price_col = None
            for col in processed_df.columns:
                if '2024-25' in str(col):
                    price_col = col
                    print(f"Found price column: '{price_col}'")
                    break
            
            if not price_col:
                print("WARNING: '2024-25' column not found! Please use Data Cleaner first to add pricing data.")
                print("Qty.MT will be set to 0.00 for all products.")
            
            if price_col and 'Assess Val.' in processed_df.columns:
                def calculate_quantity(row):
                    try:
                        price_val = row[price_col]
                        assess_val = row['Assess Val.']
                        
                        if pd.notna(price_val) and pd.notna(assess_val) and float(price_val) != 0:
                            qty = float(assess_val) / (float(price_val) * 1000)
                            return qty
                        return 0.0
                    except Exception as e:
                        return 0.0
                
                processed_df['QUANTITY_MT'] = processed_df.apply(calculate_quantity, axis=1)
                print(f"Sample quantities calculated: {processed_df['QUANTITY_MT'].head().tolist()}")
                print(f"=== END DEBUG ===\n")
            else:
                processed_df['QUANTITY_MT'] = 0.0
                print(f"=== END DEBUG ===\n")
        
        # Group by Seller PAN and Name, Buyer (NAME), and Product (HSN Desc.)
        agg_dict = {'product_value': ('VALUE', 'sum')}
        if 'QUANTITY_MT' in processed_df.columns:
            agg_dict['product_quantity'] = ('QUANTITY_MT', 'sum')
        
        seller_analysis = processed_df.groupby(['SELLER_PAN', 'SELLER_NAME', 'NAME', 'HSN Desc.']).agg(
            product_value=('VALUE', 'sum'),
            product_quantity=('QUANTITY_MT', 'sum')
        ).reset_index()
        
        # Calculate total sales per seller-buyer combination
        buyer_totals = processed_df.groupby(['SELLER_PAN', 'SELLER_NAME', 'NAME']).agg(
            buyer_total=('VALUE', 'sum'),
            buyer_quantity=('QUANTITY_MT', 'sum')
        ).reset_index()
        
        # Calculate total sales per seller
        seller_totals = processed_df.groupby(['SELLER_PAN', 'SELLER_NAME']).agg(
            seller_total=('VALUE', 'sum'),
            seller_quantity=('QUANTITY_MT', 'sum')
        ).reset_index()
        
        return {
            'seller_analysis': seller_analysis,
            'buyer_totals': buyer_totals,
            'seller_totals': seller_totals,
            'processed_data': processed_df
        }
    except Exception as e:
        print(f"Error generating seller analysis: {str(e)}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(request.url)
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected')
        return redirect(request.url)
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        input_filename = f"{timestamp}_{filename}"
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], input_filename)
        
        # Save uploaded file
        file.save(input_path)
        
        # Process the file
        output_filename = f"processed_{input_filename}"
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)
        
        success, message = process_excel_file(input_path, output_path)
        
        if success:
            # Load processed data into session
            processed_data = pd.read_excel(output_path)
            save_to_session(processed_data)
            flash(message)
            return render_template('success.html', 
                                 download_filename=output_filename,
                                 original_filename=filename)
        else:
            flash(f"Error: {message}")
            return redirect(url_for('index'))
    else:
        flash('Invalid file type. Please upload an Excel file (.xlsx or .xls)')
        return redirect(url_for('index'))

@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            flash('File not found')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'Error downloading file: {str(e)}')
        return redirect(url_for('index'))

@app.route('/summary')
def summary():
    processed_data = load_from_session()
    if processed_data is None:
        flash('No processed data found to summarize')
        return redirect(url_for('index'))

    # Generate summary
    summary_df = generate_summary(processed_data)
    if summary_df is None:
        flash('Error generating summary')
        return redirect(url_for('index'))

    # Save summary to a temporary file
    summary_file_path = os.path.join(app.config['PROCESSED_FOLDER'], 'summary_latest.xlsx')
    summary_df.to_excel(summary_file_path, index=False)

    # Pass summary to template
    return render_template('summary.html', summary=summary_df, summary_filename='summary_latest.xlsx')

@app.route('/seller_comparison')
def seller_comparison():
    processed_data = load_from_session()
    if processed_data is None:
        flash('No processed data found to analyze')
        return redirect(url_for('index'))
    
    # Check if 2024-25 column exists
    has_price_col = any('2024-25' in str(col) for col in processed_data.columns)
    if not has_price_col:
        flash('WARNING: Price column (2024-25) not found! Qty.MT will show 0.00. Please use Data Cleaner first to add pricing data.', 'warning')

    # Generate seller analysis
    analysis_data = generate_seller_analysis(processed_data)
    if analysis_data is None:
        flash('Error generating seller analysis')
        return redirect(url_for('index'))

    # Get list of unique sellers grouped by PAN (one name per PAN)
    seller_df = analysis_data['processed_data'][['SELLER_PAN', 'SELLER_NAME']].dropna()
    # Group by PAN and take the first seller name for each PAN
    unique_sellers = seller_df.groupby('SELLER_PAN')['SELLER_NAME'].first().reset_index()
    sellers = unique_sellers['SELLER_NAME'].tolist()
    sellers = [seller for seller in sellers if seller and seller != 'None']
    sellers.sort()

    return render_template('seller_comparison.html', 
                         sellers=sellers,
                         analysis_data=analysis_data)

@app.route('/compare_sellers')
def compare_sellers():
    processed_data = load_from_session()
    if processed_data is None:
        flash('No processed data found to analyze')
        return redirect(url_for('index'))

    seller1 = request.args.get('seller1')
    seller2 = request.args.get('seller2')
    page = int(request.args.get('page', 1))
    per_page = 5  # Number of buyers per page

    if not seller1 or not seller2:
        flash('Please select both sellers to compare')
        return redirect(url_for('seller_comparison'))

    # Generate seller analysis
    analysis_data = generate_seller_analysis(processed_data)
    if analysis_data is None:
        flash('Error generating seller analysis')
        return redirect(url_for('seller_comparison'))

    # Find the PAN for the selected seller names
    seller_df = analysis_data['processed_data'][['SELLER_PAN', 'SELLER_NAME']].dropna().drop_duplicates()
    seller1_pan = seller_df[seller_df['SELLER_NAME'] == seller1]['SELLER_PAN'].iloc[0] if len(seller_df[seller_df['SELLER_NAME'] == seller1]) > 0 else None
    seller2_pan = seller_df[seller_df['SELLER_NAME'] == seller2]['SELLER_PAN'].iloc[0] if len(seller_df[seller_df['SELLER_NAME'] == seller2]) > 0 else None

    # Get data for both sellers based on PAN (this will include all variations of the same company)
    seller1_data = analysis_data['processed_data'][analysis_data['processed_data']['SELLER_PAN'] == seller1_pan]
    seller2_data = analysis_data['processed_data'][analysis_data['processed_data']['SELLER_PAN'] == seller2_pan]

    # Group and sum products for each buyer (including quantity if available)
    if 'QUANTITY_MT' in seller1_data.columns:
        seller1_grouped = seller1_data.groupby(['NAME', 'HSN Desc.']).agg(
            VALUE=('VALUE', 'sum'),
            product_quantity=('QUANTITY_MT', 'sum')
        ).reset_index()
    else:
        seller1_grouped = seller1_data.groupby(['NAME', 'HSN Desc.']).agg(
            VALUE=('VALUE', 'sum')
        ).reset_index()
    
    if 'QUANTITY_MT' in seller2_data.columns:
        seller2_grouped = seller2_data.groupby(['NAME', 'HSN Desc.']).agg(
            VALUE=('VALUE', 'sum'),
            product_quantity=('QUANTITY_MT', 'sum')
        ).reset_index()
    else:
        seller2_grouped = seller2_data.groupby(['NAME', 'HSN Desc.']).agg(
            VALUE=('VALUE', 'sum')
        ).reset_index()

    # Sort by company name and product name alphabetically
    seller1_grouped = seller1_grouped.sort_values(['NAME', 'HSN Desc.'])
    seller2_grouped = seller2_grouped.sort_values(['NAME', 'HSN Desc.'])

    # Get unique buyers for each seller
    seller1_buyers_set = set(seller1_grouped['NAME'].unique())
    seller2_buyers_set = set(seller2_grouped['NAME'].unique())
    
    # Find common and unique buyers
    common_buyers = sorted(list(seller1_buyers_set.intersection(seller2_buyers_set)))
    seller1_unique_buyers = sorted(list(seller1_buyers_set - seller2_buyers_set))
    seller2_unique_buyers = sorted(list(seller2_buyers_set - seller1_buyers_set))
    
    # Create ordered buyer lists (common first, then unique)
    seller1_buyers = common_buyers + seller1_unique_buyers
    seller2_buyers = common_buyers + seller2_unique_buyers

    # Pagination for seller1
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    seller1_buyers_page = seller1_buyers[start_idx:end_idx]

    # Calculate totals
    seller1_total = seller1_data['VALUE'].sum()
    seller2_total = seller2_data['VALUE'].sum()
    
    # Calculate quantity totals if available
    seller1_qty_total = seller1_data['QUANTITY_MT'].sum() if 'QUANTITY_MT' in seller1_data.columns else 0
    seller2_qty_total = seller2_data['QUANTITY_MT'].sum() if 'QUANTITY_MT' in seller2_data.columns else 0

    # Calculate pagination info
    total_pages = (len(seller1_buyers) + per_page - 1) // per_page

    return render_template('seller_comparison_result.html',
                         seller1=seller1,
                         seller2=seller2,
                         seller1_data=seller1_grouped,
                         seller2_data=seller2_grouped,
                         seller1_buyers_page=seller1_buyers_page,
                         seller2_buyers=seller2_buyers,
                         common_buyers=common_buyers,
                         seller1_total=seller1_total,
                         seller2_total=seller2_total,
                         seller1_qty_total=seller1_qty_total,
                         seller2_qty_total=seller2_qty_total,
                         current_page=page,
                         total_pages=total_pages,
                         per_page=per_page)

@app.route('/download_analysis/<filename>')
def download_analysis(filename):
    try:
        file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            flash('Analysis file not found')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'Error downloading analysis file: {str(e)}')
        return redirect(url_for('index'))

# ========== DATA CLEANER ROUTES ==========

def clean_hsn_code(hsn):
    """Clean HSN code by removing spaces and converting to string for case-insensitive comparison"""
    if pd.isna(hsn):
        return None
    # Convert to string and remove spaces, convert to lowercase
    return str(hsn).strip().lower().replace(' ', '')

def process_data_cleaner(main_file_path, price_file_path, output_path):
    """
    Process two Excel files:
    1. Match HSN codes from both files (case-insensitive)
    2. Update HSN Desc in main file with HSN Desc from price file
    3. Add 2024-25 price column from price file to main file
    4. Only use HSN codes that appear exactly once in price file (ignore duplicates)
    5. Highlight non-updated rows in yellow
    """
    try:
        # Read both Excel files
        main_df = pd.read_excel(main_file_path)
        price_df = pd.read_excel(price_file_path)
        
        # Clean up column names - remove extra spaces and newlines
        main_df.columns = [str(col).strip().replace('\n', ' ').replace('\r', ' ') for col in main_df.columns]
        price_df.columns = [str(col).strip().replace('\n', ' ').replace('\r', ' ') for col in price_df.columns]
        
        # Find required columns
        main_hsn_col = None
        main_desc_col = None
        price_hsn_col = None
        price_desc_col = None
        price_2024_25_col = None
        
        # Search for HSN Code column in main file
        for col in main_df.columns:
            col_lower = str(col).lower().replace(' ', '').replace('_', '')
            if 'hsn' in col_lower and 'code' in col_lower:
                main_hsn_col = col
                break
        
        # Search for HSN Desc column in main file
        for col in main_df.columns:
            col_lower = str(col).lower().replace(' ', '').replace('_', '')
            if 'hsn' in col_lower and 'desc' in col_lower:
                main_desc_col = col
                break
        
        # Search for HSN Code in price file
        for col in price_df.columns:
            col_lower = str(col).lower().replace(' ', '').replace('_', '')
            if 'hsn' in col_lower and 'code' in col_lower:
                price_hsn_col = col
                break
        
        # Search for HSN Desc in price file
        for col in price_df.columns:
            col_lower = str(col).lower().replace(' ', '').replace('_', '')
            if 'hsn' in col_lower and 'desc' in col_lower:
                price_desc_col = col
                break
        
        # Search for 2024-25 column in price file
        for col in price_df.columns:
            if '2024-25' in str(col) or '2024' in str(col):
                price_2024_25_col = col
                break
        
        # Debug: Print found columns
        print(f"Main file columns: {list(main_df.columns)}")
        print(f"Price file columns: {list(price_df.columns)}")
        print(f"Found - Main HSN: {main_hsn_col}, Main Desc: {main_desc_col}")
        print(f"Found - Price HSN: {price_hsn_col}, Price Desc: {price_desc_col}, Price 2024-25: {price_2024_25_col}")
        
        # Validate required columns exist
        if not main_hsn_col:
            return False, f"HSN Code column not found in main file. Available columns: {list(main_df.columns)}", None
        if not main_desc_col:
            return False, f"HSN Desc column not found in main file. Available columns: {list(main_df.columns)}", None
        if not price_hsn_col:
            return False, f"HSN Code column not found in price file. Available columns: {list(price_df.columns)}", None
        if not price_desc_col:
            return False, f"HSN Desc column not found in price file. Available columns: {list(price_df.columns)}", None
        if not price_2024_25_col:
            return False, f"2024-25 column not found in price file. Available columns: {list(price_df.columns)}", None
        
        # Create cleaned HSN code columns for matching (case-insensitive)
        main_df['_cleaned_hsn'] = main_df[main_hsn_col].apply(clean_hsn_code)
        price_df['_cleaned_hsn'] = price_df[price_hsn_col].apply(clean_hsn_code)
        
        # First, identify duplicate HSN codes in price list (to be completely ignored)
        hsn_counts = price_df['_cleaned_hsn'].value_counts()
        duplicate_hsn = set(hsn_counts[hsn_counts > 1].index)
        
        print(f"Duplicate HSN codes found (will be ignored): {len(duplicate_hsn)}")
        print(f"Duplicate HSN list: {duplicate_hsn}")
        
        # Create a dictionary for HSN code to data mapping from price file
        # Only include HSN codes that appear exactly once (ignore duplicates completely)
        hsn_to_data = {}
        for _, row in price_df.iterrows():
            cleaned_hsn = row['_cleaned_hsn']
            # Only add if HSN is not None, not already added, and NOT a duplicate
            if cleaned_hsn and cleaned_hsn not in hsn_to_data and cleaned_hsn not in duplicate_hsn:
                hsn_to_data[cleaned_hsn] = {
                    'desc': row[price_desc_col],
                    'price': row[price_2024_25_col]
                }
        
        # Track statistics
        updated_rows = 0
        not_updated_rows = 0
        matched_hsn = len(hsn_to_data)
        
        print(f"Unique HSN codes available for matching: {matched_hsn}")
        
        # Track which rows were updated for highlighting
        updated_indices = []
        
        # Add 2024-25 column to main file if it doesn't exist
        if price_2024_25_col not in main_df.columns:
            main_df[price_2024_25_col] = None
        
        # Update HSN Desc and 2024-25 price in main file based on HSN Code matching
        for idx, row in main_df.iterrows():
            cleaned_hsn = row['_cleaned_hsn']
            if cleaned_hsn and cleaned_hsn in hsn_to_data:
                # Update HSN Desc from price file
                main_df.at[idx, main_desc_col] = hsn_to_data[cleaned_hsn]['desc']
                # Update 2024-25 price from price file
                main_df.at[idx, price_2024_25_col] = hsn_to_data[cleaned_hsn]['price']
                updated_rows += 1
                updated_indices.append(idx)
            else:
                not_updated_rows += 1
        
        # Remove temporary cleaned HSN columns
        main_df.drop(columns=['_cleaned_hsn'], inplace=True)
        
        # Calculate QTY.MT column: Assess Val. / (2024-25 * 1000)
        print("\n=== CALCULATING QTY.MT ===")
        if 'Assess Val.' in main_df.columns and price_2024_25_col in main_df.columns:
            def calculate_qty_mt(row):
                try:
                    assess_val = row['Assess Val.']
                    price = row[price_2024_25_col]
                    
                    if pd.notna(assess_val) and pd.notna(price) and float(price) != 0:
                        # QTY.MT = Assess Val. / (Price * 1000)
                        qty_mt = float(assess_val) / (float(price) * 1000)
                        return qty_mt
                    return None
                except Exception as e:
                    return None
            
            main_df['QTY.MT'] = main_df.apply(calculate_qty_mt, axis=1)
            print(f"QTY.MT column added. Sample values: {main_df['QTY.MT'].head().tolist()}")
        else:
            print("WARNING: Could not calculate QTY.MT - missing Assess Val. or price column")
        
        print("=== END QTY.MT CALCULATION ===\n")
        
        # Save to Excel with openpyxl for styling
        main_df.to_excel(output_path, index=False)
        
        # Apply yellow highlighting to non-updated rows
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Yellow fill for highlighting
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Highlight rows that were NOT updated (skip header row)
        for row_idx in range(2, len(main_df) + 2):  # Excel rows are 1-indexed, +1 for header
            df_idx = row_idx - 2  # Convert to DataFrame index
            if df_idx not in updated_indices:
                # Highlight entire row
                for cell in ws[row_idx]:
                    cell.fill = yellow_fill
        
        wb.save(output_path)
        
        stats = {
            'total_rows': len(main_df),
            'updated_rows': updated_rows,
            'not_updated_rows': not_updated_rows,
            'matched_hsn': matched_hsn
        }
        
        return True, "Files processed successfully!", stats
        
    except Exception as e:
        return False, f"Error processing files: {str(e)}", None

@app.route('/data_cleaner')
def data_cleaner():
    """Render the data cleaner upload page"""
    return render_template('data_cleaner.html')

@app.route('/data_cleaner/process', methods=['POST'])
def data_cleaner_process():
    """Process the two uploaded Excel files"""
    if 'main_file' not in request.files or 'price_file' not in request.files:
        flash('Both files are required')
        return redirect(url_for('data_cleaner'))
    
    main_file = request.files['main_file']
    price_file = request.files['price_file']
    
    if main_file.filename == '' or price_file.filename == '':
        flash('Both files must be selected')
        return redirect(url_for('data_cleaner'))
    
    if main_file and allowed_file(main_file.filename) and price_file and allowed_file(price_file.filename):
        # Save uploaded files
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        main_filename = secure_filename(main_file.filename)
        price_filename = secure_filename(price_file.filename)
        
        main_input_filename = f"{timestamp}_main_{main_filename}"
        price_input_filename = f"{timestamp}_price_{price_filename}"
        
        main_path = os.path.join(app.config['UPLOAD_FOLDER'], main_input_filename)
        price_path = os.path.join(app.config['UPLOAD_FOLDER'], price_input_filename)
        
        main_file.save(main_path)
        price_file.save(price_path)
        
        # Process the files
        output_filename = f"cleaned_{timestamp}_{main_filename}"
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)
        
        success, message, stats = process_data_cleaner(main_path, price_path, output_path)
        
        if success:
            # Update the session with the cleaned data
            processed_data = pd.read_excel(output_path)
            save_to_session(processed_data)
            flash('Data cleaned successfully! The cleaned data is now loaded for analysis.')
            
            return render_template('data_cleaner_result.html',
                                 output_filename=output_filename,
                                 total_rows=stats['total_rows'],
                                 updated_rows=stats['updated_rows'],
                                 not_updated_rows=stats['not_updated_rows'],
                                 matched_hsn=stats['matched_hsn'])
        else:
            flash(f"Error: {message}")
            return redirect(url_for('data_cleaner'))
    else:
        flash('Invalid file type. Please upload Excel files (.xlsx or .xls)')
        return redirect(url_for('data_cleaner'))

@app.route('/data_cleaner/download/<filename>')
def data_cleaner_download(filename):
    """Download the processed cleaned file"""
    try:
        file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            flash('File not found')
            return redirect(url_for('data_cleaner'))
    except Exception as e:
        flash(f'Error downloading file: {str(e)}')
        return redirect(url_for('data_cleaner'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)